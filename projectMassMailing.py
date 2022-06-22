import openpyxl as xl
import smtplib 
def sep():
    print("-"*50)

def send_mail(content,to, step, maxx):
    try:
        s = smtplib.SMTP(f'smtp.outlook.com', 587)
        s.starttls()
        s.login("hritiksingh_21186@aitpune.edu.in", "********")
        message = "\n" + content
        s.sendmail("hritiksingh_21186@aitpune.edu.in", to, message)
        if step == maxx:
            s.quit()
    except Exception as e:
        print("Mail Error", "Kindly check the login credentials.\n\nIf using GMail, we are counted among the less secure apps and for some users our services might not be available.\n\nWe are sorry for the inconvinence caused.")
        pass
    
def create_mail(name, v1, txt):
    return txt.replace(v1, name)
    
def extract_info():
    path = input("Enter the path of the sheet:\n")
    sep()
    text = input("Enter the body of the EMail:\n")
    sep()
    var = input("Enter the variable used:\n")
    sep()
    obj = xl.load_workbook(path).active
    for i in range(2, obj.max_row+1):
        name = obj.cell(row=i, column=2).value
        email = obj.cell(row=i, column=3).value
        interest = obj.cell(row=i, column=4).value
        body = create_mail(name, var, text)
        send_mail(body, email, i, obj.max_row + 1)
        pass
    
def main():
    extract_info()
    pass

main()